#!/usr/bin/env python3

import os
import sys
import re
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# ✅ 讓 script 可以跨平台使用相對路徑
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from utils.file_utils import find_latest_test_folder, get_spec_json_path_by_product

# ✅ 自動設定根目錄與 Spec 資料夾路徑
base_folder = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
spec_folder = os.path.join(base_folder, "spec_reference")
family_mapping_file = os.path.join(spec_folder, "family_mapping.json")

# ---------- 比對 & 分析 ----------
def analyze_results(csv_path, spec_json_path):
    try:
        df = pd.read_csv(csv_path)
    except Exception as e:
        print(f"❌ 無法讀取 CSV 檔案: {e}")
        return

    try:
        with open(spec_json_path, "r") as f:
            spec_data = json.load(f)
    except Exception as e:
        print(f"❌ 無法讀取 SPEC JSON 檔案: {e}")
        return

    product_name = os.path.basename(csv_path).split("_fio_summary_results.csv")[0]
    family_key = product_name.split("-")[0] + "-" + product_name.split("-")[1]

    raw_capacity = product_name.split("-")[-1].replace("TB", "")
    capacity = f"{float(raw_capacity):.2f}TB"
    print(f"📌 判斷出的容量名稱為: {capacity}")

    def parse_tb_string(tb_str):
        return float(tb_str.replace("TB", "").strip())

    capacity_index = -1
    if family_key in spec_data:
        available_caps = spec_data[family_key]["Capacity"]
        cap_diffs = [abs(parse_tb_string(c) - parse_tb_string(capacity)) for c in available_caps]
        min_diff = min(cap_diffs)
        if min_diff <= 0.5:
            capacity_index = cap_diffs.index(min_diff)
            print(f"📌 自動對應容量: {available_caps[capacity_index]}")
        else:
            print(f"❌ 找不到接近 {capacity} 的容量（誤差 > 0.5TB）")
            return
    else:
        print(f"❌ 找不到型號 {family_key} 的 spec")
        return

    def check_pass(row):
        name = row["Test Name"]
        iops = row["IOPS"]
        if "128KB_Seq_Read" in name:
            spec_val = spec_data[family_key]["128KB Seq Read (MB/s)"][capacity_index]
            actual = float(row["Bandwidth"].replace("MB/s", ""))
        elif "128KB_Seq_Write" in name:
            spec_val = spec_data[family_key]["128KB Seq Write (MB/s)"][capacity_index]
            actual = float(row["Bandwidth"].replace("MB/s", ""))
        elif "Random_Read" in name and "16KB" in name:
            spec_val = spec_data[family_key]["16KB Random Read (KIOPs)"][capacity_index]
            actual = iops / 1000
        elif "Random_Write" in name and "16KB" in name:
            spec_val = spec_data[family_key]["16KB Random Write (KIOPs)"][capacity_index]
            actual = iops / 1000
        elif "RandRW" in name and "16KB" in name:
            spec_val = spec_data[family_key]["16KB Random Mixed 70/30 RR/RW (KIOPs)"][capacity_index]
            actual = iops / 1000
        elif "Random_Read" in name and "4KB" in name:
            spec_val = spec_data[family_key]["4KB Random Read (KIOPs)"][capacity_index]
            actual = iops / 1000
        elif "Random_Write" in name and "4KB" in name:
            spec_val = spec_data[family_key]["4KB Random Write (KIOPs)"][capacity_index]
            actual = iops / 1000
        elif "RandRW" in name and "4KB" in name:
            spec_val = spec_data[family_key]["4KB Random Mixed 70/30 RR/RW (KIOPs)"][capacity_index]
            actual = iops / 1000
        else:
            return pd.Series(["N/A", "GRAY", None])

        lower_bound = spec_val * 0.9
        if actual >= spec_val:
            result = "PASS"
            color = "GREEN"
        elif lower_bound <= actual < spec_val:
            result = "+/-10% PASS"
            color = "YELLOW"
        else:
            result = "FAIL"
            color = "RED"

        return pd.Series([result, color, spec_val])

    df[["Result", "Color", "Spec Value"]] = df.apply(lambda row: check_pass(row), axis=1)

    output_path = csv_path.replace(".csv", "_analyzed.xlsx")
    df.drop(columns=["Color"]).to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    fill_colors = {
        "GREEN": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "RED": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "YELLOW": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "GRAY": PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    }

    for row in range(2, ws.max_row + 1):
        color = df.loc[row - 2, "Color"]
        if color in fill_colors:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill_colors[color]

    wb.save(output_path)
    print(f"✅ 分析報告已儲存到: {output_path}")

# ---------- 主流程 ----------
def main():
    latest_folder = find_latest_test_folder()
    if not latest_folder:
        return

    csv_files = [f for f in os.listdir(latest_folder) if f.endswith("_fio_summary_results.csv")]
    if not csv_files:
        print("❌ 找不到測試結果 CSV")
        return

    print("📄 找到以下測試結果:")
    for idx, name in enumerate(csv_files):
        print(f"[{idx}] {name}")
    csv_choice = int(input("請選擇要分析的 CSV 檔案編號: "))
    csv_path = os.path.join(latest_folder, csv_files[csv_choice])

    product_name = os.path.basename(csv_path).split("_fio_summary_results.csv")[0]
    spec_path = get_spec_json_path_by_product(product_name)
    if not spec_path:
        return

    analyze_results(csv_path, spec_path)

if __name__ == "__main__":
    main()
