#!/usr/bin/env python3

import logging
import subprocess
import os
from datetime import datetime
import openpyxl

# Execute command and return output
def run_command(command):
    try:
        result = subprocess.run(command, shell=True, capture_output=True, text=True, check=True)
        return result.stdout
    except subprocess.CalledProcessError:
        return ""
    
def setup_logging(log_file):
    logging.basicConfig(
        filename=log_file,
        filemode="w",
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s"
    )
    # 同時輸出到終端
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    console.setFormatter(formatter)
    logging.getLogger("").addHandler(console)

def set_cpu_frequency_performance():
    """
    使用 cpupower 將 CPU frequency 設定為 performance 模式，並回傳每個 CPU 核心的設定結果。
    """
    result = []
    try:
        # 檢查 cpupower 是否已安裝
        if run_command("which cpupower").strip() == "":
            print("cpupower not found. Installing...")
            subprocess.run("sudo apt install -y linux-tools-common linux-tools-$(uname -r)", shell=True, check=True)

        # 設定所有 CPU frequency governor 為 performance
        print("Setting CPU frequency governor to performance...")
        subprocess.run("sudo cpupower frequency-set -g performance", shell=True, check=True)

        # 驗證每個 CPU 的 governor 狀態
        cpu_count = os.cpu_count()
        for cpu in range(cpu_count):
            governor = run_command(f"cat /sys/devices/system/cpu/cpu{cpu}/cpufreq/scaling_governor").strip()
            if governor == "performance":
                result.append(f"cpu{cpu}: performance")
            else:
                result.append(f"cpu{cpu}: failed to set performance")
    except subprocess.CalledProcessError as e:
        print(f"Error setting CPU frequency governor: {e}")
        result.append("Error: Unable to set performance mode.")
    return result

def get_power_state(device):
    """
    根據 NVMe 裝置名稱（如 nvme0），讀取其 power_state。
    Args:
        device: NVMe 基本裝置名稱 (例如: nvme0, nvme1)
    Returns:
        power_state: 讀取到的 Power State，若失敗則回傳 "Unknown"
    """
    try:
        power_state_path = f"/sys/class/nvme/{device}/device/power_state"
        if os.path.exists(power_state_path):
            with open(power_state_path, "r") as f:
                return f.read().strip()
        else:
            return "Path does not exist"
    except Exception as e:
        print(f"Error retrieving power state for {device}: {e}")
    return "Unknown"

def get_nvme_devices():
    """Get NVMe devices and their capacities."""
    command = ["lshw", "-c", "storage", "-businfo"]
    try:
        result = subprocess.run(command, capture_output=True, text=True, check=True)
        output = result.stdout
    except subprocess.CalledProcessError as e:
        print("Error running lshw:", e)
        return []

    rows = []
    for line in output.splitlines()[2:]:
        row = line.split(None, 3)
        if len(row) >= 3 and "nvme" in row[1]:
            bus_info = row[0]
            device = row[1].replace("/dev/", "")  # 清理設備名稱
            description = row[3] if len(row) > 3 else "Unknown"
            capacity = get_device_capacity(device)  # 獲取設備容量
            rows.append([bus_info, device, description, capacity])
    return rows

def get_device_capacity(device):
    """Get NVMe device capacity using lsblk."""
    try:
        # 加上 n1 確保匹配完整設備名稱
        full_device = f"{device}n1" if not device.endswith("n1") else device
        
        # 執行 lsblk，提取設備名稱和容量
        output = run_command(f"lsblk -b -o NAME,SIZE")
        for line in output.splitlines():
            columns = line.split()
            if full_device in columns[0]:  # 匹配設備名稱
                size = int(columns[1])
                return f"{size / 1e12:.2f} TB" if size >= 1e12 else f"{size / 1e9:.2f} GB"
    except Exception as e:
        print(f"Error retrieving capacity for {device}: {e}")
    return "Unknown"


def get_nvme_firmware():
    """Get NVMe firmware versions using nvme-cli."""
    firmware_info = {}
    try:
        command = ["nvme", "list"]
        result = subprocess.run(command, capture_output=True, text=True, check=True)
        for line in result.stdout.splitlines()[2:]:  # 忽略標題行
            columns = line.split()
            if len(columns) >= 8:
                device_full = columns[0].replace("/dev/", "")  # 提取 nvme0n1
                firmware = columns[-1]  # Firmware 版本
                # 同時保存 nvme0 和 nvme0n1 的映射
                base_device_name = device_full.split("n")[0]
                firmware_info[base_device_name] = firmware
                firmware_info[device_full] = firmware
        return firmware_info
    except subprocess.CalledProcessError as e:
        print(f"Error retrieving firmware versions: {e}")
    return firmware_info


# Function to get temperature thresholds and current temperature using smartctl
def get_temperature_thresholds(device):
    """Get temperature thresholds and current temperature using smartctl."""
    try:
        full_device = f"/dev/{device}"  # 確保完整路徑名稱
        result = subprocess.run(["sudo", "smartctl", "-a", full_device], capture_output=True, text=True, check=True)
        warning, critical, current_temp = "Unknown", "Unknown", "Unknown"
        for line in result.stdout.splitlines():
            if "Warning  Comp. Temp. Threshold:" in line:
                warning = line.split(":")[-1].strip()
            elif "Critical Comp. Temp. Threshold:" in line:
                critical = line.split(":")[-1].strip()
            elif "Temperature:" in line and "Celsius" in line:
                current_temp = line.split(":")[-1].strip().split()[0] + " Celsius"
        return warning, critical, current_temp
    except subprocess.CalledProcessError as e:
        print(f"Error retrieving temperature for {device}: {e}")
        return "Error", "Error", "Error"
    
def enable_io_polling(base_device):
    """
    Enable I/O polling for the specified device.

    Args:
        base_device (str): The device name (e.g., 'nvme0n1').

    Returns:
        str: A message indicating the result of enabling I/O polling.
    """
    io_poll_path = f"/sys/block/{base_device}/queue/io_poll"
    io_poll_result = ""
    logging.info(f"Enabling I/O polling for {base_device}...")
    if os.path.exists(io_poll_path):
        try:
            subprocess.run(f"echo 1 | sudo tee {io_poll_path} > /dev/null", 
                           shell=True, check=True, stderr=subprocess.DEVNULL)
            with open(io_poll_path, "r") as f:
                current_value = f.read().strip()
                if current_value == "1":
                    logging.info(f"I/O polling successfully set to 1 for {base_device}.")
                    io_poll_result = f"I/O polling successfully set to 1 for {base_device}."
                else:
                    logging.warning(f"I/O polling not supported for {base_device}.")
                    io_poll_result = f"I/O polling not supported for {base_device}."
        except subprocess.CalledProcessError:
            logging.warning(f"I/O polling not supported for {base_device}.")
            io_poll_result = f"I/O polling command failed for {base_device}."
    else:
        logging.warning(f"I/O polling path {io_poll_path} does not exist.")
        io_poll_result = f"I/O polling path {io_poll_path} does not exist."
    
    return io_poll_result

def enable_poll_queues():
    """Enable poll_queues by reloading nvme module and setting parameter."""
    poll_queue_value = "Unknown"
    try:
        # 執行 modprobe
        logging.info("Reloading nvme module with poll_queues=4...")
        subprocess.run("sudo modprobe -r nvme && sudo modprobe nvme poll_queues=4", shell=True, check=True)

        # 獲取 poll_queues 的值
        poll_queues_path = "/sys/module/nvme/parameters/poll_queues"
        if os.path.exists(poll_queues_path):
            with open(poll_queues_path, "r") as f:
                poll_queue_value = f.read().strip()
                logging.info(f"poll_queues value: {poll_queue_value}")
        else:
            logging.warning(f"{poll_queues_path} does not exist.")
    except subprocess.CalledProcessError as e:
        logging.error(f"Failed to reload nvme module with poll_queues=4: {e}")
    except Exception as e:
        logging.error(f"Error reading poll_queues value: {e}")
    
    return poll_queue_value


def create_excel_report(folder_name, devices_list, governors, power_states, selected_devices, firmware_info, io_poll_results, poll_queue_value):
    """Generate the Excel report."""
    filename = os.path.join(folder_name, "SSD_testing_list.xlsx")
    wb = openpyxl.Workbook()

    # Sheet 1: SSD list
    sheet1 = wb.active
    sheet1.title = "SSD list"
    sheet1.append(["Bus info", "Device", "Description", "Capacity"])
    for row in devices_list:
        sheet1.append(row)

    # Sheet 2: SSD status
    sheet2 = wb.create_sheet(title="SSD status")
    sheet2.append([
        "Bus info", "Device", "Firmware", "Description",
        "Warning Temp Threshold", "Critical Temp Threshold", "Current Temperature"
    ])
    for bus_info, device, description, capacity in devices_list:
        if device in selected_devices:
            base_device_name = f"{device}n1"  # 確保使用 nvme0n1
            firmware = firmware_info.get(device, firmware_info.get(base_device_name, "Unknown"))
            warning, critical, current_temp = get_temperature_thresholds(base_device_name)
            sheet2.append([bus_info, device, firmware, description, warning, critical, current_temp])

    # Sheet 3: Env_version
    sheet3 = wb.create_sheet(title="Env_version")
    sheet3.append(["Tool", "Version"])
    tools = ["nvme", "smartctl", "lspci", "fio", "python3"]
    for tool in tools:
        version = run_command(f"{tool} --version || {tool} -V || {tool} -v").splitlines()
        sheet3.append([tool, version[0] if version else "Not Installed"])

    linux_version = run_command("cat /etc/os-release | grep PRETTY_NAME | cut -d= -f2").strip().strip('"')
    sheet3.append(["Linux Version", linux_version if linux_version else "Unknown"])

    kernel_version = run_command("uname -r").strip()
    sheet3.append(["Kernel Version", kernel_version if kernel_version else "Unknown"])

    # Sheet 4: Testing_Env
    sheet4 = wb.create_sheet(title="Testing_Env")
    
    # CPU Governor 設定
    sheet4.append(["CPU Governor"])
    for line in governors:  # governors 包含每個 CPU 的設定狀態
        sheet4.append([line])
    
    # Selected NVMe 和 Power State
    sheet4.append([])
    sheet4.append(["Selected NVMe Devices"])
    for device, power_state in zip(selected_devices, power_states):
        sheet4.append([f"{device} - Power State: {power_state}"])
    sheet4.append([])
    for device, io_poll_result in zip(selected_devices, io_poll_results):
        sheet4.append([f"I/O Polling Result for {device}: {io_poll_result}"])
    sheet4.append([])
    sheet4.append([f"poll_queues Value: {poll_queue_value}"])
   
    wb.save(filename)
    print(f"Excel report saved to: {filename}")


def main():
    # 建立時間戳和輸出資料夾
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    folder_name = f"./Solidigm_Testing_Result_{current_time}"
    os.makedirs(folder_name, exist_ok=True)

    # 設定日誌輸出檔案
    log_file_path = os.path.join(folder_name, "script.log")
    setup_logging(log_file_path)

    logging.info("Script started.")

    try:
        # 設定 CPU 頻率
        logging.info("Setting CPU frequency governor to performance...")
        governors = set_cpu_frequency_performance()
        logging.info("CPU governors set successfully.")

        # 獲取 NVMe 設備資訊
        devices_list = get_nvme_devices()
        logging.info(f"Retrieved NVMe devices: {devices_list}")

        # 獲取 Firmware 資訊
        firmware_info = get_nvme_firmware()
        logging.info("Firmware information retrieved successfully.")

        # 顯示可用設備
        logging.info("\nAvailable NVMe Devices:")
        for idx, (_, device, description, capacity) in enumerate(devices_list):
            logging.info(f"{idx}: /dev/{device} - {description} ({capacity})")

        # 多選設備
        selected_indices = input("\nEnter the indices of the devices to test (comma-separated, e.g., 0,1,2): ")
        selected_indices = [int(idx.strip()) for idx in selected_indices.split(",")]
        selected_devices = [devices_list[idx][1] for idx in selected_indices]
        logging.info(f"Selected devices: {', '.join(selected_devices)}")

        power_states = []
        io_poll_results = []

        # 針對每個選中的設備執行操作
        for selected_device in selected_devices:
            try:
                base_device = f"{selected_device}n1" if not selected_device.endswith("n1") else selected_device
                logging.info(f"Processing device: {base_device}")

                # 生成 SMART log
                log_file_path = os.path.join(folder_name, f"{base_device}_before_testing_smartctl_log.txt")
                logging.info(f"Generating SMART log for {base_device}...")
                with open(log_file_path, "w") as log_file:
                    subprocess.run(["sudo", "smartctl", "-a", f"/dev/{base_device}"], stdout=log_file, check=True)
                logging.info(f"SMART log saved to: {log_file_path}")

                # 啟用 I/O polling
                io_poll_result = enable_io_polling(base_device)
                io_poll_results.append(io_poll_result)
                logging.info(f"I/O polling result for {base_device}: {io_poll_result}")

                # 啟用 poll_queues
                poll_queue_value = enable_poll_queues()
                logging.info(f"poll_queues Value for {base_device}: {poll_queue_value}")

                # 讀取 SSD Power State
                power_state = get_power_state(selected_device)
                power_states.append(power_state)
                logging.info(f"Power state for {selected_device}: {power_state}")

            except Exception as e:
                logging.error(f"An error occurred while processing {selected_device}: {e}")
                power_states.append("Error")
                io_poll_results.append("Error")

        # 產生 Excel 報告
        create_excel_report(folder_name, devices_list, governors, power_states, selected_devices, firmware_info, io_poll_results, poll_queue_value)

        logging.info("Excel report generated successfully.")

    except Exception as e:
        logging.error(f"An error occurred: {e}")
    finally:
        logging.info("Script finished.")

if __name__ == "__main__":
    main()